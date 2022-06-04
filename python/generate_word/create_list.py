import json
from pathlib import Path
import os
import pysrt
import sys
from jisho_api.tokenize import Tokens
from jisho_api.word import Word
from openpyxl import Workbook
import requests
from requests.utils import quote

nameFile = sys.argv[1]
nameExcel = nameFile.replace(".srt", ".xlsx")
sub_a = pysrt.open(nameFile)
arr = []
current_dich = 0
max = 20
arr = [sub_a[i*max:(i+1)*max] for i in range(len(sub_a)//max)]
last = sub_a[len(arr)*max:len(sub_a)]
arr.append(last)
max_line = len(sub_a)
print(f'=======================')
print(f'File {nameFile} có {max_line} câu thoại')
print(f'=======================')

ROOT = Path.home() / ".jisho/data/mazii"
ROOT_2 = Path.home() / ".jisho/data/tokens"
def phan_tich_cau(index=0):
    print('Chạy phân tích câu')
    is_infinity = True
    count = index * max + 1
    while is_infinity:
        for x in range(index, len(arr)):
            for y in arr[x]:
                text = y.text.replace(u'\u3000', u' ').replace('\n', ' ')
                currentRow = (count, str(y.start), text)
                ws.append(currentRow)
                if (ROOT_2 / (text + ".json")).exists():
                    with open(ROOT_2 / (text + ".json"), "r") as fp:
                        result = json.load(fp)
                        if (result is not None):
                            for res in result['data']:
                                if (res['pos_tag'] == 'Noun'):
                                    ws.append(('', '', '', str(res['token']), 'Danh từ'))
                                elif (res['pos_tag'] == 'Verb'):
                                    ws.append(('', '', '', str(res['token']), 'Động từ'))
                else:
                    result = Tokens.request(text)
                    if (result is not None):
                        result = json.loads(result.json())
                        for res in result['data']:
                            if (res['pos_tag'] == 'Noun'):
                                ws.append(('', '', '', str(res['token']), 'Danh từ'))
                            elif (res['pos_tag'] == 'Verb'):
                                ws.append(('', '', '', str(res['token']), 'Động từ'))
                        save2(text, result)
                count += 1
                print('[Phân tích] Hàng: ' + str(x) +' || ' + str(count - 1) + '/' + str(max_line))
                if (count % 10 == 0):
                    print('[Phân tích] Save !!')
                    wb.save(filename=nameExcel)
        wb.save(filename=nameExcel)
        if (count >= max_line):
            print('Xong phân tích câu')
            is_infinity = False

def dich_jisho(index=2):
    print('Chạy từ_điển_1')
    count_0 = 0
    for row in ws.iter_rows(min_row=index):
        count_0 += 1
        print('[từ_điển_1] Hàng: ', str(row[0].row))
        if (row[3].value is not None):
            result = Word.request(row[3].value, True)
            if (result is not None):
                row[5].value = str(result.data[0].japanese[0].word)
        if (count_0 % 20 == 0):
            wb.save(filename=nameExcel)
    wb.save(filename=nameExcel)
    print('Xong từ_điển_1')

def dich_tu(index=2):
    print('Chạy mazii')
    count_1 = 0
    for row in ws.iter_rows(min_row=index):
        count_1 += 1
        if ((row[5].value is not None) and (row[5].value != 'None')):
            word = row[5].value
            print('[mazii] Hàng: ' + str(row[0].row) + ' ' + str(row[5].value))
            if (ROOT / (word + ".json")).exists():
                with open(ROOT / (word + ".json"), "r") as fp:
                    json_res = json.load(fp)
                    if (json_res.get('found')):
                        row[6].value = json_res.get('data')[0].get('means')[0].get('mean')
                        row[5].value = json_res.get('data')[0].get('phonetic')
            else:
                response = requests.get('https://mazii.net/api/search/' + quote(row[5].value) + '/10/1')
                json_res = response.json()
                if (json_res.get('found')):
                    row[6].value = json_res.get('data')[0].get('means')[0].get('mean')
                    row[5].value = json_res.get('data')[0].get('phonetic')
                save(word, json_res)
        elif (row[3].value is not None):
            word = row[3].value
            print('[mazii] Hàng: ' + str(row[0].row) + ' ' + str(row[3].value))
            if (ROOT / (word + ".json")).exists():
                with open(ROOT / (word + ".json"), "r") as fp:
                    json_res = json.load(fp)
                    if (json_res.get('found')):
                        row[6].value = json_res.get('data')[0].get('means')[0].get('mean')
                        row[5].value = json_res.get('data')[0].get('phonetic')
            else:
                response = requests.get('https://mazii.net/api/search/' + quote(row[3].value) + '/10/1')
                json_res = response.json()
                if (json_res.get('found')):
                    row[6].value = json_res.get('data')[0].get('means')[0].get('mean')
                    row[5].value = json_res.get('data')[0].get('phonetic')
                save(word, json_res)

        if (count_1 % 20 == 0):
            wb.save(filename=nameExcel)

    wb.save(filename=nameExcel)
    print('Xong mazii')
    

def save(word, r):
    ROOT.mkdir(exist_ok=True)
    with open(ROOT / f"{word}.json", "w") as fp:
        fp.write(json.dumps(r, indent=4, ensure_ascii=False))
        
def save2(word, r):
    ROOT.mkdir(exist_ok=True)
    with open(ROOT_2 / f"{word}.json", "w") as fp:
        fp.write(json.dumps(r, indent=4, ensure_ascii=False))

def main():
    try:
        os.makedirs(ROOT)    
        print("Directory " , ROOT ,  " Created ")
    except FileExistsError:
        print("Directory " , ROOT ,  " already exists")
    try:
        os.makedirs(ROOT_2)    
        print("Directory " , ROOT_2 ,  " Created ")
    except FileExistsError:
        print("Directory " , ROOT_2 ,  " already exists")  

    global wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Từ vựng"
    rowTitle = ('STT', 'Thời gian', 'Subtitle', 'Từ vựng', 'Loại từ', 'Từ điền', 'Nghĩa')
    ws.append(rowTitle)
    phan_tich_cau()
    dich_jisho()
    dich_tu()
    wb.save(filename=nameExcel)
    print('\nThành công!!!\nFile Xuất >>>>> ' + nameExcel )

if __name__ == "__main__":
    main()